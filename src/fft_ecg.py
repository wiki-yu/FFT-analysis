from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import os
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.patches import FancyArrowPatch
from mpl_toolkits.mplot3d import proj3d
from scipy.fftpack import fft
import operator
from scipy import signal
from scipy.stats import kstest,ks_2samp
from statsmodels.distributions.empirical_distribution import ECDF

def oneD_fft(y,z):
    N = len(y)
    print(N)
    # sample spacing
    #T = 0.001
    T = 0.001
    base = np.linspace(0.0, N*T, N)
 
    y = y - np.mean(y)
    z = z - np.mean(z)
    yf = fft(y)
    zf = fft(z)
    xf = np.linspace(0.0, 1.0/(2.0*T), N/2)
    
    plt.figure()
    plt.subplot(211)
    yfft_val = 2.0/N * np.abs(yf[0:N/2])
    zfft_val = 2.0/N * np.abs(zf[0:N/2])
    
    ks_2samp(yfft_val, zfft_val)
    a = yfft_val
    b = zfft_val
    ecdf1, ecdf2 = ECDF(a), ECDF(b)
    fig = plt.figure(1)
    xs = np.linspace(min(a+b),max(a+b), num=10000)
    #plt.figure(figsize=(12,8))
    plt.plot(xs,ecdf1(xs), xs,ecdf2(xs))
    plt.show()
    print(ks_2samp(a,b))
    
    
    
    
    fig = plt.figure(2)
    plt.subplot(211)
    plt.plot(xf, yfft_val,color='r', label='Abl distal electrode')
    plt.legend(loc="upper right")
    plt.ylabel('Magnitude')
    plt.grid(True)
    #plt.title("Spectral graph")
    plt.subplot(212)
    plt.plot(xf, zfft_val,color='b',label='Cs proximal electrode')
    plt.legend(loc="upper right")
    plt.xlabel('Frequency(HZ)')
    plt.ylabel('Magnitude')
    plt.grid(True)
    plt.show()
    

def read_2d_data():
    wb = load_workbook('point16.xlsx')
    sheet_1 = wb.get_sheet_by_name('point')
 
    x = np.zeros(sheet_1.max_row-1) 
    y = np.zeros(sheet_1.max_row-1) 
    z = np.zeros(sheet_1.max_row-1) 
 
    for i in range(1,sheet_1.max_row):
        x[i-1]=sheet_1.cell(row=i+1, column=13).value #abl p
        y[i-1]=sheet_1.cell(row=i+1, column=14).value #abl d
        z[i-1]=sheet_1.cell(row=i+1, column=5).value #cs p
    return x,y,z

x,y,z = read_2d_data()
xx = x[:2048:2]
yy = y[:2048:2]
zz = z[:2048:2]
oneD_fft(yy,zz)


base = np.arange(1,len(xx)+1,1) 

fig = plt.figure(3)
#plt.title('Electrogram of Abl p, Abl d and Cs p')
plt.plot(base,xx,'g-',label = 'Abl proximal electrode')
plt.plot(base,yy,'r-',label = 'Abl distal electrode')
plt.plot(base,zz,'b-',label = 'Cs proximal electrode')
plt.legend(loc='upper right')
plt.grid(True)
plt.xlabel('Time stamp seq')
plt.ylabel('Magnitude')
plt.show()


fig = plt.figure(4)
#plt.subplot(311)
#plt.title('Electrogram of Abl p, Abl d and Cs p')
#plt.plot(base,xx,'g-',label = 'Abl p')
#plt.ylim([-2.0,1.5])
#plt.grid()
#plt.ylabel('Abl p')
plt.subplot(211)
plt.plot(base,yy,'r-',label = 'Abl distal electrode')
plt.legend(loc='upper right')

plt.grid(True)
plt.subplot(212)
plt.plot(base,zz,'b-',label = 'Cs proximal electrode')
plt.legend(loc='upper right')
plt.ylabel('Magnitude')
plt.xlabel('Time stamp seq')
plt.grid(True)
plt.show()


